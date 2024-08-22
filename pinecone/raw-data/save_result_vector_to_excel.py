import os
from openpyxl import Workbook, load_workbook
from typing import List, Dict, Callable

def save_to_excel(keyword: str, results: List[Dict], base_filename: str, find_excel_file_func: Callable):
    file_path = find_excel_file_func(base_filename)

    if file_path:
        print(f"Tìm thấy file tại: {file_path}")
    else:
        print(f"Không tìm thấy file {base_filename}. Sẽ tạo file mới trong thư mục hiện tại.")
        file_path = base_filename

    attempt = 1
    while True:
        try:
            if os.path.exists(file_path):
                wb = load_workbook(file_path)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Search Results"
                ws.append(["Keyword", "Score", "Content"])

            for result in results:
                metadata = result['metadata']
                content = metadata.get('text', 'Không có nội dung')

                ws.append([
                    keyword,
                    result['score'],
                    content
                ])

            wb.save(file_path)
            print(f"Kết quả đã được thêm vào file: {file_path}")
            break
        except PermissionError:
            print(f"Không thể ghi vào file {file_path}. Đang thử tạo file mới...")
            attempt += 1
            file_path = f"search_results_pinecone_with_raw_data_{attempt}.xlsx"
        except Exception as e:
            print(f"Lỗi khi lưu file: {e}")
            break